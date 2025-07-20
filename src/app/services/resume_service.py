import os
import re
import fitz
from flask import current_app
from src.app.utils import file_service_util
import pytesseract
from PIL import Image
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from src.app.ai_engine.task import retrieve_resume_data


class ResumeService:
    
    def get_porcessed_resume_data(self):
        extract_text = ""
        upload_file = os.path.join(current_app.config['UPLOAD_FOLDER'], "AKK.pdf")
        ext = file_service_util.get_file_extension(upload_file)
        
        try:
            if ext in ('jpg', 'jpeg', 'png'):
                extract_text = self._process_image_file(upload_file)
            elif ext == 'pdf':
                extract_text = self._process_pdf_file(upload_file)
            elif ext == 'docx':
                extract_text = self._process_docx_file(upload_file)
            elif ext == 'xlsx':
                extract_text = self._process_excel_file(upload_file)
            
            if extract_text:
                resume_data = retrieve_resume_data(extract_text, "gemini")
                return resume_data
                
        except Exception as ex:
            current_app.logger.error(f"Error in get_porcessed_resume_data: {ex}")
            return None
    
    def _process_image_file(self, file_path) -> str:
        image = Image.open(file_path)
        image_text = pytesseract.image_to_string(image, lang='eng+mya+tha')
        return self._extract_text_from_image(image_text)
    
    def _process_pdf_file(self, file_path) -> str:
        extract_text = ""
        pdf_document = fitz.open(file_path)
        for page in pdf_document:
            extract_text += page.get_text()
        return extract_text
    
    def _process_docx_file(self, file_path) -> str:
        docx_text = self._extract_text_from_docx(file_path)
        if docx_text is not None:
            return self._remove_duplicate_text_from_doc(docx_text)
        return ""
    
    def _process_excel_file(self, file_path) -> str:
        xlsx_text = []
        workbook = load_workbook(file_path)
        
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    if cell.value is not None and cell.data_type != 'f':
                        cell_data = str(cell.value) if cell.value is not None else ""
                        row_data.append(cell_data)
                xlsx_text.append(" ".join(row_data))
            xlsx_text.append(" ")
            
        return "\n".join(xlsx_text)
    
    def _extract_text_from_image(self, text) -> str:
        clean_text = ""
        clean_text = re.sub(r'[^a-zA-Z0-9\s.,@/\u1000-\u109F\u0E00-\u0E7F]', '', text)
        clean_text = re.sub(r'\n+', '\n', clean_text)
        clean_text = "\n".join([line.strip() for line in clean_text.split('\n')])
        return clean_text
    
    def _extract_text_from_docx(self, docx_file) -> str | None:
        try:
            document = Document(docx_file)
            full_text = []

            for block in document.iter_inner_content():
                if isinstance(block, Paragraph):
                    if block.text.strip():
                        full_text.append(block.text.strip())
                elif isinstance(block, Table):
                    table_content = []
                    for row in block.rows:
                        row_data = []
                        for cell in row.cells:
                            cell_text = cell.text.strip()
                            row_data.append(cell_text)
                        table_content.append(" | ".join(row_data))
                    full_text.append("\n".join(table_content))

            return "\n".join(full_text)
            
        except Exception as ex:
            current_app.logger.error(f"Error during text extract for docx file: {ex}")
            raise ValueError(f"Error during text extract for docx file: {ex}")
    
    def _remove_duplicate_text_from_doc(self, raw_text) -> str:
        lines = raw_text.split('\n')
        processed_lines = []

        for line in lines:
            if '|' in line:
                parts = line.split('|')
                unique_parts = []
                seen_parts = set()
                
                for part in parts:
                    stripped_part = part.strip()
                    if stripped_part not in seen_parts:
                        seen_parts.add(stripped_part)
                        unique_parts.append(stripped_part)
                        
                processed_line = ' | '.join(unique_parts)
                processed_lines.append(processed_line)
            else:
                processed_lines.append(line)
                
        return '\n'.join(processed_lines)

resume_service = ResumeService()